import openpyxl, os
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from io import BytesIO

path = r'D:\xwechat_files\wxid_72expwi31maq22_4f6a\msg\file\2026-06'
pics_dir = r'D:\Users\21365\Pictures'
src = os.path.join(path, '申请材料清单（企业开发者）.xlsx')
wb = openpyxl.load_workbook(src)
ws = wb.worksheets[1]

thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
bf = Font(size=9)
wa = Alignment(wrap_text=True, vertical='top')

# ── Row 4: 心率 ──
ws.cell(row=4, column=1, value='心率数据读取权限\n(HEART_RATE)\n\n华为账号OAuth 2.0授权\n用户主动同意后获取').font=bf
ws.cell(row=4, column=1).alignment=wa; ws.cell(row=4, column=1).border=thin

ws.cell(row=4, column=2, value=(
    '【数据使用场景】\n'
    '1. 首页健康仪表盘：展示从华为手表同步的实时心率(bpm)，用户可查看当前心率和变化趋势。\n'
    '2. 健康记录详情页：展示今日/本周/本月心率曲线图，含平均心率、最大/最小心率统计。\n'
    '3. 运动安全监测：用户记录运动时结合心率和血糖综合分析，心率过高时提醒注意安全。\n'
    '4. AI预测增强：心率作为TCN模型第14维特征(f14)，提升血糖预测精度。\n'
    '   心率升高→交感神经兴奋→升糖激素分泌→血糖升高。系统学习个体心率-血糖关联。\n'
    '5. 异常心率提醒：静息心率>100bpm或<50bpm时提示用户关注心血管健康。\n'
    '6. 长期趋势分析：周/月报中展示心率与血糖关联图表，帮助发现规律。'
)).font=bf; ws.cell(row=4, column=2).alignment=wa; ws.cell(row=4, column=2).border=thin

ws.cell(row=4, column=3, value=(
    '【数据使用目的】\n'
    '1. 健康数据综合展示：用户在一个App内查看血糖+心率+步数+睡眠，无需切换多个App。\n'
    '2. 运动安全防护：糖尿病患者运动时易低血糖，实时心率配合血糖可提前预警。\n'
    '   参考：运动心率不应超过(220-年龄)×70%。\n'
    '3. 血糖预测增强：f14=最近1小时心率均值，是TCN深度学习15维输入特征之一。\n'
    '4. 心血管健康监测：长期心率数据趋势反映自主神经功能，与糖尿病并发症相关。\n'
    '5. 个性化建议生成：SmartAdvisor引擎综合心率+血糖+运动数据生成个性化建议。\n'
    '6. 科研统计（脱敏）：匿名化心率统计数据用于算法优化和糖尿病管理研究。'
)).font=bf; ws.cell(row=4, column=3).alignment=wa; ws.cell(row=4, column=3).border=thin

ws.cell(row=4, column=4, value=(
    '【数据读取时机】\n'
    '1. 首次授权后立即读取近7天历史心率（约10080条记录）。\n'
    '2. 前台服务每30分钟读取最新心率（最近1小时窗口，约60条）。\n'
    '3. 用户首页下拉或进入健康记录页时实时读取。\n'
    '4. 运动期间每5分钟高频读取以监测运动强度。\n'
    '5. 日终23:59读取当日完整数据用于日报生成。\n'
    '6. 所有读取的数据存储在本地Room数据库，不上传云端。\n'
    '7. 用户可在设置中清除所有心率数据。'
)).font=bf; ws.cell(row=4, column=4).alignment=wa; ws.cell(row=4, column=4).border=thin

ws.cell(row=4, column=5, value='【截图】\n设置页-数据来源\n（显示华为手表\n 心率/步数/睡眠接入）').font=bf
ws.cell(row=4, column=5).alignment=wa; ws.cell(row=4, column=5).border=thin

ws.cell(row=4, column=6, value=(
    '1. 心率数据仅存本地，不上传云端\n'
    '2. 仅用于上述用途，不用于广告/画像\n'
    '3. 用户可随时导出或删除\n'
    '4. 传输用HTTPS/TLS 1.3加密\n'
    '5. 遵循《个人信息保护法》相关要求\n'
    '6. App无后台静默读取，需前台服务运行\n'
    '7. 用户关闭Health Kit授权后立即停止读取'
)).font=bf; ws.cell(row=4, column=6).alignment=wa; ws.cell(row=4, column=6).border=thin

# ── Row 5: 步数 ──
ws.cell(row=5, column=1, value='步数数据读取权限\n(STEPS)\n\n华为账号OAuth 2.0授权\n用户主动同意后获取').font=bf
ws.cell(row=5, column=1).alignment=wa; ws.cell(row=5, column=1).border=thin

ws.cell(row=5, column=2, value=(
    '【数据使用场景】\n'
    '1. 首页仪表盘：显示当日实时步数及目标完成进度条（默认8000步/天，可自定义）。\n'
    '2. 健康记录页：展示7天/30天步数趋势图，含日均步数、活跃天数统计。\n'
    '3. 运动记录关联：用户记录运动时自动关联步数，计算消耗热量和降糖效果。\n'
    '4. 活动量评估报告：周/月报展示步数与TIR(目标范围内时间)的关联分析。\n'
    '5. AI预测增强：步数作为TCN模型第15维特征(f15)，评估活动量对血糖的影响。'
)).font=bf; ws.cell(row=5, column=2).alignment=wa; ws.cell(row=5, column=2).border=thin

ws.cell(row=5, column=3, value=(
    '【数据使用目的】\n'
    '1. 日常活动量监测：ADA建议糖尿病患者每日步行30分钟（约4000-5000步）。\n'
    '2. 运动降糖量化：系统分析步数-血糖关联，计算"每千步降糖量"。\n'
    '3. AI模型特征：f15=最近1小时步数总和。久坐→低步数→胰岛素敏感度↓→血糖↑。\n'
    '4. 个性化运动建议：SmartAdvisor基于步数+血糖生成运动建议。\n'
    '5. DallaMan参数调整：活动水平基于步数动态调整k1参数(非胰岛素依赖利用)。'
)).font=bf; ws.cell(row=5, column=3).alignment=wa; ws.cell(row=5, column=3).border=thin

ws.cell(row=5, column=4, value=(
    '【数据读取时机】\n'
    '1. 首次授权后读取近7天历史步数（每日汇总+每小时明细）。\n'
    '2. 前台服务每30分钟读取最新步数（与血糖同步频率）。\n'
    '3. 用户首页下拉或进入健康记录页时实时读取。\n'
    '4. 日终23:59读取当日完整步数用于日报。\n'
    '5. 所有数据存本地，不上传云端。'
)).font=bf; ws.cell(row=5, column=4).alignment=wa; ws.cell(row=5, column=4).border=thin

ws.cell(row=5, column=5, value='【截图】\n首页健康仪表盘\n（展示当日步数\n统计和血糖趋势）').font=bf
ws.cell(row=5, column=5).alignment=wa; ws.cell(row=5, column=5).border=thin

ws.cell(row=5, column=6, value=(
    '1. 步数数据仅存本地，不上传云端\n'
    '2. 仅用于上述用途，不用于广告/画像\n'
    '3. 用户可随时导出或删除\n'
    '4. 遵循数据最小化原则\n'
    '5. App无后台静默读取'
)).font=bf; ws.cell(row=5, column=6).alignment=wa; ws.cell(row=5, column=6).border=thin

# ── Row 6: 睡眠 ──
ws.cell(row=6, column=1, value='睡眠数据读取权限\n(SLEEP)\n\n华为账号OAuth 2.0授权\n用户主动同意后获取').font=bf
ws.cell(row=6, column=1).alignment=wa; ws.cell(row=6, column=1).border=thin

ws.cell(row=6, column=2, value=(
    '【数据使用场景】\n'
    '1. 健康记录睡眠页：展示睡眠时长（深睡/浅睡/REM）、睡眠评分、入睡/醒来时间。\n'
    '2. 夜间血糖关联分析：睡眠数据与血糖曲线叠加展示，直观看到睡眠-血糖关系。\n'
    '3. 黎明现象判断：结合醒来时间和清晨血糖(4-8点升高)判断黎明现象。\n'
    '4. 夜间低血糖评估：NightMonitor模块结合睡眠时段预测夜间低血糖风险。'
)).font=bf; ws.cell(row=6, column=2).alignment=wa; ws.cell(row=6, column=2).border=thin

ws.cell(row=6, column=3, value=(
    '【数据使用目的】\n'
    '1. 睡眠-血糖关联：睡眠<6小时→次日胰岛素抵抗+30%。系统追踪个人规律。\n'
    '2. 黎明现象识别：约50%T1DM患者存在。结合醒来时间+晨间升幅自动识别。\n'
    '3. 夜间低血糖预警：睡眠时段为高风险窗口，睡前提前提醒补充碳水。\n'
    '4. 作息规律建议：通过睡眠规律性分析提供改善建议，帮助稳定血糖。\n'
    '5. ★睡眠数据不参与AI预测模型计算（仅用于趋势分析），保证模型稳定性。'
)).font=bf; ws.cell(row=6, column=3).alignment=wa; ws.cell(row=6, column=3).border=thin

ws.cell(row=6, column=4, value=(
    '【数据读取时机】\n'
    '1. 首次授权后读取近7天历史睡眠（每日汇总记录）。\n'
    '2. 每天早上8:00自动读取前一晚完整睡眠数据。\n'
    '3. 用户进入健康记录-睡眠页面时实时读取。\n'
    '4. 读取频率极低（每日1次），对功耗和流量影响极小。\n'
    '5. 所有数据存本地，不上传云端。'
)).font=bf; ws.cell(row=6, column=4).alignment=wa; ws.cell(row=6, column=4).border=thin

ws.cell(row=6, column=5, value='【截图】\n设置页-数据来源\n（显示华为手表\n睡眠数据接入）').font=bf
ws.cell(row=6, column=5).alignment=wa; ws.cell(row=6, column=5).border=thin

ws.cell(row=6, column=6, value=(
    '1. 睡眠数据仅存本地，不上传云端\n'
    '2. 不参与AI预测模型实时计算\n'
    '3. 读取频率低(日1次)，功耗影响极小\n'
    '4. 用户可随时导出或删除'
)).font=bf; ws.cell(row=6, column=6).alignment=wa; ws.cell(row=6, column=6).border=thin

# ── Row 7: 体重 ──
ws.cell(row=7, column=1, value='体重数据写入权限\n(WEIGHT)\n\n华为账号OAuth 2.0授权\n用户主动同意后获取').font=bf
ws.cell(row=7, column=1).alignment=wa; ws.cell(row=7, column=1).border=thin

ws.cell(row=7, column=2, value=(
    '【数据使用场景】\n'
    '1. 设置页体重录入：用户在"我的→身高体重"手动输入体重(kg)，系统自动同步到Health Kit。\n'
    '2. 体重趋势追踪：健康记录页展示7/30/90天体重变化趋势图。\n'
    '3. 预测模型个性化：体重直接用于DallaMan七隔室生理模型的核心参数计算。\n'
    '4. 胰岛素剂量参考：体重变化影响ISF和CR，系统更新体重后自动重算个性化参数。'
)).font=bf; ws.cell(row=7, column=2).alignment=wa; ws.cell(row=7, column=2).border=thin

ws.cell(row=7, column=3, value=(
    '【数据使用目的（核心：血糖预测模型个性化）】\n'
    '1. 葡萄糖分布体积: Vg=体重×1.6 dL/kg(范围60-300)\n'
    '   → 体重越大→分布体积越大→同等碳水升糖幅度越小\n'
    '2. 胰岛素分布体积: Vi=体重×0.05 L/kg(范围2-25)\n'
    '   → 体重越大→胰岛素分布越广→同等剂量降糖越弱\n'
    '3. 基础胰岛素Ib和基础血糖Gb的个性化设定\n'
    '4. ISF自动估算: TDD=日均总剂量(与体重正相关)\n'
    '5. 写入Health Kit目的: 跨App数据互通，避免用户重复录入'
)).font=bf; ws.cell(row=7, column=3).alignment=wa; ws.cell(row=7, column=3).border=thin

ws.cell(row=7, column=4, value=(
    '【数据写入时机】\n'
    '1. ★仅在用户主动输入体重并点击"保存"后写入。\n'
    '2. ★无任何后台自动写入行为。\n'
    '3. 写入前显示确认信息:"体重将同步到华为Health Kit"。\n'
    '4. 写入频率极低(用户通常每周更新1-2次)。\n'
    '5. 体重数据同时存本地和Health Kit云端。'
)).font=bf; ws.cell(row=7, column=4).alignment=wa; ws.cell(row=7, column=4).border=thin

ws.cell(row=7, column=5, value='【截图】\n设置页-身高体重\n输入界面\n（显示"用于预测\n模型个性化参数计算"）').font=bf
ws.cell(row=7, column=5).alignment=wa; ws.cell(row=7, column=5).border=thin

ws.cell(row=7, column=6, value=(
    '1. 仅用户主动触发写入，无后台行为\n'
    '2. 写入数据仅含体重值和记录时间\n'
    '3. 用户可在Health Kit中随时删除\n'
    '4. 传输存储用HTTPS/AES加密\n'
    '5. 用于医疗目的(糖尿病管理)\n'
    '6. 遵循《个人信息保护法》'
)).font=bf; ws.cell(row=7, column=6).alignment=wa; ws.cell(row=7, column=6).border=thin

# Column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 52
ws.column_dimensions['C'].width = 48
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 30

# Row heights
for r in range(4, 8):
    ws.row_dimensions[r].height = 350

# Insert screenshots
screenshot_map = {
    4: '微信图片_20260616204954_379_44.jpg',   # 心率/睡眠: settings data sources
    5: '微信图片_20260616204952_376_44.jpg',   # 步数: home page
    6: '微信图片_20260616204954_379_44.jpg',   # 睡眠: settings data sources
    7: '微信图片_20260616204955_380_44.jpg',   # 体重: body info
}

for row, pic_name in screenshot_map.items():
    pic_path = os.path.join(pics_dir, pic_name)
    if not os.path.exists(pic_path):
        print(f'MISSING: {pic_path}')
        continue
    pil_img = PILImage.open(pic_path)
    ratio = 200.0 / pil_img.size[0]
    new_h = int(pil_img.size[1] * ratio)
    pil_img = pil_img.resize((200, new_h), PILImage.LANCZOS)
    buf = BytesIO()
    pil_img.save(buf, format='JPEG', quality=90)
    buf.seek(0)
    xl_img = XLImage(buf)
    xl_img.width = 200
    xl_img.height = new_h
    xl_img.anchor = f'E{row}'
    ws.add_image(xl_img)
    print(f'Row {row}: {pic_name} ({200}x{new_h})')

out = os.path.join(path, '申请材料清单（企业开发者）_糖盾.xlsx')
wb.save(out)
print(f'\nDone: {out}')
